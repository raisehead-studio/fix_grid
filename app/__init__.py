import re
import sqlite3
import os

from datetime import datetime, timedelta
from flask import Flask, render_template, redirect, request, url_for, flash, abort, send_file, jsonify, current_app
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from jinja2 import TemplateNotFound
from werkzeug.security import check_password_hash, generate_password_hash
from .models import get_user_by_username, get_user_by_id_with_role, get_role_page_permissions_from_db
from .utils import page_name_map, check_page_view_permission
from io import BytesIO
from openpyxl import load_workbook

# routes
from .role_routes import role_bp
from .account_routes import account_bp
from .power_routes import power_bp
from .water_routes import water_bp
from .taiwater_power_routes import taiwater_power_bp
from .disaster_routes import disaster_bp

login_manager = LoginManager()

def needs_password_update(password_updated_at, user_id):
    # 沒有更新時間或超過 90 天
    if not password_updated_at:
        return True
    try:
        last_updated = datetime.fromisoformat(password_updated_at)
        if datetime.now() - last_updated > timedelta(days=90):
            return True
    except Exception:
        return True

    # 沒有登入紀錄
    try:
        conn = sqlite3.connect('kao_power_water.db', timeout=10)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM user_login_logs WHERE user_id = ?", (user_id,))
        count = c.fetchone()[0]
        return count == 0
    except Exception as e:
        print("DB error in needs_password_update:", e)
        return True  # fallback: require password update
    finally:
        if 'conn' in locals():
            conn.close()

def validate_password(pwd):
    if len(pwd) < 12:
        return False
    if not re.search(r'[A-Z]', pwd):
        return False
    if not re.search(r'[a-z]', pwd):
        return False
    if not re.search(r'\d', pwd):
        return False
    if not re.search(r'[^A-Za-z0-9]', pwd):  # 特殊符號
        return False
    return True

def create_app():
    app = Flask(__name__)  # 移除內建靜態檔案服務
    app.secret_key = 'supersecretkey'
    
    # 設定會話過期時間（例如：8小時）
    from datetime import timedelta
    app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
    
    # 設定會話為永久性（但會在指定時間後過期）
    app.config['SESSION_PERMANENT'] = True

    login_manager.init_app(app)
    login_manager.login_view = 'login'
    app.register_blueprint(role_bp)
    app.register_blueprint(account_bp)
    app.register_blueprint(power_bp)
    app.register_blueprint(water_bp)
    app.register_blueprint(taiwater_power_bp)
    app.register_blueprint(disaster_bp)

    # DNS 驗證路由 - 不需要登入驗證
    @app.route('/.well-known/<path:filename>')
    def dns_verification(filename):
        """處理 DNS 驗證檔案，例如 /.well-known/acme-challenge/xxx"""
        # 安全性檢查：防止路徑遍歷攻擊
        if '..' in filename or filename.startswith('/'):
            abort(404)
        
        # 只允許特定副檔名
        allowed_extensions = {'.txt', '.html', '.htm', '.json'}
        if not any(filename.endswith(ext) for ext in allowed_extensions):
            abort(404)
            
        try:
            return send_file(f'static/.well-known/{filename}')
        except FileNotFoundError:
            abort(404)

    # @app.route('/static/<path:filename>')
    # @login_required
    # def static_files(filename):
    #     """處理靜態檔案，需要登入驗證"""
    #     # 安全性檢查：防止路徑遍歷攻擊
    #     if '..' in filename or filename.startswith('/'):
    #         abort(404)

    #     # 只允許特定檔案類型（不包含 Excel 檔案）
    #     allowed_extensions = {'.js', '.css', '.png', '.jpg', '.jpeg', '.gif', '.ico', '.svg'}
    #     if not any(filename.endswith(ext) for ext in allowed_extensions):
    #         abort(404)

    #     try:
    #         return send_file(f'static/{filename}')
    #     except FileNotFoundError:
    #         abort(404)

    # @app.route('/<path:filename>')
    # def public_files(filename):
    #     """處理其他公開檔案，例如 /robots.txt, /sitemap.xml 等"""
    #     # 安全性檢查：防止路徑遍歷攻擊
    #     if '..' in filename or filename.startswith('/') or filename.startswith('static/'):
    #         abort(404)
        
    #     # 只允許特定檔案類型
    #     allowed_extensions = {'.txt', '.xml', '.json', '.html'}
    #     if not any(filename.endswith(ext) for ext in allowed_extensions):
    #         abort(404)
            
    #     # 只允許特定檔案名稱（白名單）
    #     allowed_files = {'robots.txt', 'sitemap.xml', 'favicon.ico'}
    #     if not any(filename == allowed_file for allowed_file in allowed_files):
    #         abort(404)
            
    #     try:
    #         return send_file(f'static/{filename}')
    #     except FileNotFoundError:
    #         abort(404)

    @app.before_request
    def block_options():
        if request.method == 'OPTIONS':
            abort(403)

    class User(UserMixin):
        def __init__(self, id, username, full_name, phone, district_id, district, village_id, village, role_id, role_name, password_updated_at, ip=None):
            self.id = id
            self.username = username
            self.full_name = full_name
            self.phone = phone
            self.district_id = district_id
            self.district = district
            self.village_id = village_id
            self.village = village
            self.role_id = role_id
            self.role_name = role_name
            self.password_updated_at = password_updated_at
            self.ip = ip

    def insert_login_log(user_id, ip):
        try:
            conn = sqlite3.connect('kao_power_water.db', timeout=10)
            c = conn.cursor()
            c.execute('''
                INSERT INTO user_login_logs (user_id, ip, login_time)
                VALUES (?, ?, ?)
            ''', (user_id, ip, datetime.now()))
            conn.commit()
        except Exception as e:
            print("DB error in insert_login_log:", e)
        finally:
            if 'conn' in locals():
                conn.close()

    @login_manager.user_loader
    def load_user(user_id):
        user = get_user_by_id_with_role(user_id)
        if user:
            return User(
                user['id'], user['username'], user['full_name'], user['phone'],
                user['district_id'], user['district'], user['village_id'], user['village'], user['role_id'], user['role_name'], user['password_updated_at']
            )
        return None

    @app.route('/', methods=['GET', 'POST'])
    def login():
        # 如果用戶已經登入，自動跳轉到 profile 頁面
        if current_user.is_authenticated:
            return redirect(url_for('page_info', page='profile'))

        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            user = get_user_by_username(username)
            if user and check_password_hash(user['password'], password):
                ip = request.remote_addr

                login_user(User(
                    user['id'], user['username'], user['full_name'], user['phone'],
                    user['district_id'], user['district'], user['village_id'], user['village'], user['role_id'], user['role_name'], user['password_updated_at'], ip=ip
                ), remember=True)

                if needs_password_update(user['password_updated_at'], user['id']):
                    print('change')
                    return redirect(url_for('force_change_password'))

                # 記錄登入紀錄
                insert_login_log(user['id'], ip)

                return redirect(url_for('page_info', page='profile'))
            else:
                flash("帳號或密碼錯誤", "danger")
        return render_template('login.html')

    @app.route('/force_change_password', methods=['GET', 'POST'])
    @login_required
    def force_change_password():
        if request.method == 'POST':
            ip = request.remote_addr
            new_password = request.form['new_password']

            try:
                # 取得原本的加密密碼
                conn = sqlite3.connect('kao_power_water.db', timeout=10)
                c = conn.cursor()
                c.execute("SELECT password FROM users WHERE id = ?", (current_user.id,))
                row = c.fetchone()
                old_password_hash = row[0] if row else None

                # 檢查格式與一致性
                if not validate_password(new_password):
                    flash("密碼格式不符：需至少12碼，包含大小寫英文字母、數字、特殊符號", "danger")
                    return render_template('force_change_password.html')

                if old_password_hash and check_password_hash(old_password_hash, new_password):
                    flash("新密碼不得與舊密碼相同", "danger")
                    return render_template('force_change_password.html')

                hashed = generate_password_hash(new_password)
                c.execute("UPDATE users SET password = ?, password_updated_at = ? WHERE id = ?", (hashed, datetime.now().isoformat(), current_user.id))
                conn.commit()
            except Exception as e:
                print("DB error in force_change_password:", e)
                flash("更新密碼失敗", "danger")
                return render_template('force_change_password.html')
            finally:
                if 'conn' in locals():
                    conn.close()

            insert_login_log(current_user.id, ip)

            flash("密碼更新成功，請重新登入", "success")
            logout_user()
            return redirect(url_for('login'))

        return render_template('force_change_password.html')

    @app.route('/dashboard')
    @login_required
    def dashboard():
        return f"Hello {current_user.full_name}, you're logged in with role name {current_user.role_name}."

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    @app.route("/api/login_logs/<int:user_id>")
    @login_required
    def get_login_logs_by_user(user_id):
        try:
            conn = sqlite3.connect("kao_power_water.db", timeout=10)
            c = conn.cursor()
            c.execute("""
                SELECT ip, login_time
                FROM user_login_logs
                WHERE user_id = ?
                ORDER BY login_time DESC
                LIMIT 100
            """, (user_id,))
            logs = []
            for row in c.fetchall():
                ip, raw_time = row
                try:
                    parsed_time = datetime.fromisoformat(raw_time)
                    formatted_time = parsed_time.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    formatted_time = raw_time
                logs.append({"ip": ip, "login_time": formatted_time})
            return {"logs": logs}
        except Exception as e:
            print("DB error in /api/login_logs:", e)
            return {"error": "資料庫錯誤"}, 500
        finally:
            if 'conn' in locals():
                conn.close()
        return {"logs": logs}

    @app.route('/page/<page>')
    @login_required
    def page_info(page):
        # 權限檢查
        all_permissions = get_role_page_permissions_from_db()
        role_name = current_user.role_name
        
        # 檢查頁面查看權限（需要 view 權限）
        if not check_page_view_permission(page, role_name, all_permissions):
            abort(403)  # 權限不足
        
        role_pages = all_permissions.get(role_name, {})
        actions = role_pages.get(page, [])

        # 將通用變數整理
        context = {
            "role": role_name,
            "current_page": page,
            "current_page_name": page_name_map.get(page, page),
            "actions": actions,
            "pages": role_pages,
            "page_name_map": page_name_map,
            "user_permissions": actions,
            "current_user": current_user,
        }

        if current_user.role_id == 4:  # 里幹事
            context["page_name_map"]["power_outage"] = "停電通報（表一）"
            context["page_name_map"]["water_outage"] = "停水通報（表二）"
        elif current_user.role_id in [5, 6]:  # 台電/台水人員
            context["page_name_map"]["power_outage"] = "停電彙整（表一）"
            context["page_name_map"]["water_outage"] = "停水彙整（表二）"
        else:
            context["page_name_map"]["power_outage"] = "停電彙整表（表一）"
            context["page_name_map"]["water_outage"] = "停水彙整表（表二）"

        # 預設檢查該頁面是否有對應模板
        target_template = f"{page}.html"

        try:
            return render_template(target_template, **context)
        except TemplateNotFound:
            return render_template("page_info.html", **context)

    @app.route("/api/export-excel", methods=["POST"])
    def export_excel():
        try:
            payload = request.get_json()
            template_name = payload.get("template")
            filename = payload.get("filename")
            data = payload.get("data")
            start_row = payload.get("start_row", 1)
            start_col = payload.get("start_col", 1)

            if not template_name or not filename or not isinstance(data, list):
                return jsonify({"error": "Missing or invalid parameters"}), 400

            template_path = os.path.join(current_app.root_path, "static", template_name)
            if not os.path.exists(template_path):
                return jsonify({"error": f"Template file '{template_name}' not found"}), 404

            wb = load_workbook(template_path)
            ws = wb.active

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

            # 民國時間格式產出
            now = datetime.now()
            roc_year = now.year - 1911
            time_string = f"統計時間：{roc_year}年{now.month}月{now.day}日（{now.hour}:{now.minute:02}）"

            # 根據範本名稱設定對應欄位
            if template_name == "sheet4.xlsx":
                ws["I1"] = time_string
            elif template_name == "sheet5.xlsx":
                ws["E1"] = time_string
            elif template_name == "sheet6.xlsx":
                ws["H1"] = time_string

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name=filename,
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.before_request
    def attach_ip_to_current_user():
        if current_user.is_authenticated:
            current_user.ip = request.headers.get("X-Forwarded-For", request.remote_addr)

    return app

if __name__ == "__main__":
    app.run(debug=True)
